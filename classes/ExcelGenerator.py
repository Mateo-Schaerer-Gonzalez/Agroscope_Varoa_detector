"""Excel generation utilities for the Varroa detector project."""

import os
import cv2
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage


class ExcelGenerator:
    """Handles Excel file generation with data and images."""
    
    def __init__(self):
        # Define color fills
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def create_recordings_summary(self, summary_df, time_survival_path, excel_path, 
                                recordings_base_path):
        """Create Excel summary with recordings data."""
        # Guard clause
        if summary_df.empty:
            raise ValueError("No summary data available for Excel generation")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        recordings = sorted(summary_df['recording'].unique())
        
        # Add overview sheet
        ws = wb.create_sheet(title="Overview")
        self._add_image_to_sheet(ws, time_survival_path, 'A1')
        
        # Add individual recording sheets
        for rec in recordings:
            self._create_recording_sheet(wb, summary_df, rec, recordings_base_path)
        
        wb.save(excel_path)
        print(f"Summary Excel with sheets per recording saved to {excel_path}")
    
    def create_mites_summary(self, mite_data, zones, by_mite_path, excel_path, 
                           frame_path, stage_zones):
        """Create Excel summary with mite status data."""
        # Guard clause
        if mite_data.empty:
            raise ValueError("No mite data available for Excel generation")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        for zone in zones:
            self._create_mite_zone_sheet(wb, mite_data, zone, by_mite_path, 
                                       frame_path, stage_zones)
        
        wb.save(excel_path)
        print(f"Excel summary with zone-wise sheets saved to {excel_path}")
    
    def _add_image_to_sheet(self, ws, image_path, anchor):
        """Add an image to a worksheet at the specified anchor."""
        try:
            if os.path.exists(image_path):
                img = XLImage(image_path)
                img.anchor = anchor
                ws.add_image(img)
        except Exception as e:
            print(f"Failed to add image {image_path}: {e}")
    
    def _create_recording_sheet(self, wb, summary_df, recording_num, base_path):
        """Create a sheet for a specific recording."""
        rec_df = summary_df[summary_df['recording'] == recording_num]
        
        ws = wb.create_sheet(title=f"Recording {recording_num}")
        
        # Add data
        for r_idx, row in enumerate(dataframe_to_rows(rec_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Add images
        survival_path = os.path.join(base_path, f'recording{recording_num}', 'survival.png')
        self._add_image_to_sheet(ws, survival_path, 'H1')
        
        detection_path = os.path.join(base_path, f'recording{recording_num}', 'frame_0.jpg')
        try:
            if os.path.exists(detection_path):
                img = XLImage(detection_path)
                img.width = img.width * 0.3
                img.height = img.height * 0.3
                img.anchor = 'H31'
                ws.add_image(img)
        except Exception as e:
            print(f"Failed to add detection image: {e}")
    
    def _create_mite_zone_sheet(self, wb, mite_data, zone, by_mite_path, 
                              frame_path, stage_zones):
        """Create a sheet for a specific mite zone."""
        zone_df = mite_data[mite_data['zone ID'] == zone]
        
        # Pivot so rows = mite ID, cols = recording, values = status
        pivot = zone_df.pivot(index='mite ID', columns='recording', values='status')
        
        ws = wb.create_sheet(title=str(zone))
        
        # Write headers and data with conditional formatting
        for r_idx, row in enumerate(dataframe_to_rows(pivot.reset_index(), 
                                                     index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 1 and c_idx > 1:  # Only data cells (skip headers)
                    if value == 'alive':
                        cell.fill = self.green_fill
                    elif value == 'dead':
                        cell.fill = self.red_fill
        
        # Add zone plot image
        zone_plot_path = os.path.join(by_mite_path, "zones", f"{zone}.png")
        self._add_zone_image(ws, zone_plot_path)
        
        # Add detection label and ROI images
        self._add_roi_images(ws, zone, stage_zones, frame_path, by_mite_path)
    
    def _add_zone_image(self, ws, image_path):
        """Add zone plot image to worksheet."""
        if os.path.exists(image_path):
            try:
                img = XLImage(image_path)
                img.width = img.width * 0.5
                img.height = img.height * 0.5
                img.anchor = "H2"
                ws.add_image(img)
            except Exception as e:
                print(f"Failed to add zone image: {e}")
    
    def _add_roi_images(self, ws, zone, stage_zones, frame_path, by_mite_path):
        """Add ROI images for the zone."""
        # Add header for ROI section
        ws.merge_cells('S1:U1')
        ws['S1'] = "first recording detections"
        ws['S1'].font = ws['S1'].font.copy(bold=True)
        ws['S1'].alignment = ws['S1'].alignment.copy(horizontal='center')
        
        # Add ROI images
        try:
            frame0 = cv2.imread(frame_path)
            if frame0 is None:
                print(f"Frame 0 image not found at {frame_path}")
                return
            
            idx = 0
            for zone_obj in stage_zones:
                if any(z.text == zone for z in zone_obj.text_zones):
                    idx += 1
                    self._create_and_add_roi_image(ws, zone_obj, frame0, zone, 
                                                 idx, by_mite_path)
                    
        except Exception as e:
            print(f"Failed to add ROI images: {e}")
    
    def _create_and_add_roi_image(self, ws, zone_obj, frame0, zone, idx, by_mite_path):
        """Create and add a single ROI image."""
        try:
            img = zone_obj.get_ROI(frame0)
            img = cv2.resize(img, (0, 0), fx=0.5, fy=0.5)
            
            roi_img_path = os.path.join(by_mite_path, "zones", f"{zone}_roi{idx * 5}.jpg")
            cv2.imwrite(roi_img_path, img)
            
            roi_excel_img = XLImage(roi_img_path)
            roi_excel_img.anchor = f"S{1 + idx * 5}"
            ws.add_image(roi_excel_img)
            
        except Exception as e:
            print(f"Failed to create ROI image for zone {zone}: {e}")
